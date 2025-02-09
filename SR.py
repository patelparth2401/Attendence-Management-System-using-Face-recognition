from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image,ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
def Student_registration():
    bg="#06283D"
    framebg="#EDEDED"
    framefg="#06283D"
    
    root=Tk()
    root.title("student registraiton")
    root.geometry("1200x600+0+0")
    root.config(bg=bg)
    
    
    file=pathlib.Path('studentdata.xlsx')
    if file.exists():
        pass
    else:
        file=Workbook()
        sheet=file.active
        sheet['A1']="Enrollment No."
        sheet['B1']="Name"
        sheet['C1']="Class"
        sheet['D1']="Gender"
        sheet['E1']="DOB"
        sheet['F1']="Date of registration"
        sheet['G1']="Father Name"
        sheet['H1']="Mother Name"
        
        file.save('studentdata.xlsx')
    
    #gender
    def selection():
        global gender
        value=radio.get()
        if value==1:
            gender="Male"
        else:
            gender="female"
            
    
    '''*********************************show image*****************************************   '''     
    def showimg():
        global filename
        global img
        filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="select image file",
                                            filetype=(("JPEG file","*.jpg"),("PNG files","*.png"),("ALL files","*.txt")))
       
        img=(Image.open(filename))
        resized_img=img.resize((190,190))
        photo2=ImageTk.PhotoImage(resized_img)
        lbl.config(image=photo2)
        lbl.image=photo2
    '''******************************************Exit*************************************** '''
    def Exit():
        root.destroy()
    '''**********************************************clear********************************************'''
    def clear():
        Name.set('')
        DOB.set('')
        Class.set('Select class')
        Registration.set('')
        F_Name.set('')
        M_Name.set("")
        imageicon7=Image.open("user.jpg")
        imageicon7=imageicon7.resize((200,200))
        imageicon7=ImageTk.PhotoImage(imageicon7)
        lbl.config(image=imageicon7)
        lbl.image=imageicon7
        
    "*********************************************************8save****************************************"
    def save():
        
        R1=Registration.get()
        N1=Name.get()
        C1=Class.get()
        try:
            G1=gender
        except:
            messagebox.showerror("error","Select Gender!")
        D1=DOB.get()
        D2=Date.get()
        F1=F_Name.get()
        M1=M_Name.get()
        
        if N1=='' or R1=='' or C1=='Select Class' or D1=='' or F1=='' or M1=='':
            messagebox.showerror("eroor","Few values are missing")
        else:
            file=openpyxl.load_workbook("studentdata.xlsx")
            sheet=file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=R1)
            sheet.cell(column=2,row=sheet.max_row,value=N1)
            sheet.cell(column=3,row=sheet.max_row,value=C1)
            sheet.cell(column=4,row=sheet.max_row,value=G1)
            sheet.cell(column=5,row=sheet.max_row,value=D1)
            sheet.cell(column=6,row=sheet.max_row,value=D2)
            sheet.cell(column=7,row=sheet.max_row,value=F1)
            sheet.cell(column=8,row=sheet.max_row,value=M1)
            file.save(r'studentdata.xlsx')
            
            try:
                img.save("Student image/"+str(R1)+".jpg")
            except:
                messagebox.showinfo("info","picture is not available!")
            messagebox.showinfo('info',"Sucessfully data entered!")
            
            clear()
            
    "***********************************************Search*********************************************"
    def search():
        text=Search.get()
        clear()
        file=openpyxl.load_workbook("studentdata.xlsx")
        sheet=file.active
        for row in sheet.rows:
            if row[0].value == text:
                name=row[0]
                er_no_pos=str(name)[14:-1]
                er_number=str(name)[15:-1]
                
                
        try:
            print(str(name))
        except:
            messagebox.showerror('Invalid',"Invalid number")
            
        x1=sheet.cell(row=int(er_number),column=1).value
        x2=sheet.cell(row=int(er_number),column=2).value
        x3=sheet.cell(row=int(er_number),column=3).value
        x4=sheet.cell(row=int(er_number),column=4).value
        x5=sheet.cell(row=int(er_number),column=5).value
        x6=sheet.cell(row=int(er_number),column=6).value
        x7=sheet.cell(row=int(er_number),column=7).value
        x8=sheet.cell(row=int(er_number),column=8).value
        
        Registration.set(x1)
        Name.set(x2)
        Class.set(x3)
        
        if x4=="Female":
            R2.select()
        else:
            R1.select()
        DOB.set(x5)
        Date.set(x6)
        F_Name.set(x7)
        M_Name.set(x8)
        
        imageicon7=Image.open("Student image/"+str(x1)+".jpg")
        imageicon7=imageicon7.resize((200,200))
        imageicon7=ImageTk.PhotoImage(imageicon7)
        
        lbl.config(image=imageicon7)
        lbl.image=imageicon7
        
    "************************************Update******************************************"
    def update():
        
        R1=Registration.get()
        N1=Name.get()
        C1=Class.get()
        selection()
        G1=gender
        D1=DOB.get()
        D2=Date.get()
        F1=F_Name.get()
        M1=M_Name.get()
        
        file=openpyxl.load_workbook("studentdata.xlsx")
        sheet=file.active
        for row in sheet.rows:
            if row[0].value==R1:
                name=row[0]
                print(str(name))
                er_no_pos=str(name)[14:-1]
                er_number=str(name)[15:-1]
                
            
        sheet.cell(column=1,row=int(er_number),value=R1)
        sheet.cell(column=2,row=int(er_number),value=N1)
        sheet.cell(column=3,row=int(er_number),value=C1)
        sheet.cell(column=4,row=int(er_number),value=G1)
        sheet.cell(column=5,row=int(er_number),value=D1)
        sheet.cell(column=6,row=int(er_number),value=D2)
        sheet.cell(column=7,row=int(er_number),value=F1)
        sheet.cell(column=8,row=int(er_number),value=M1)
        
        file.save(r'studentdata.xlsx')
        
        try:
            img.save("Student image/"+str(x1)+".jpg")
        except:
            pass
    
        messagebox.showinfo("info","Updated sucessfully!!")
        clear()
                 
            
            
            
    #top frames
    Label(root,text="Email:12002080701042@gmail.com",width=10,height=3,bg="#f0687c",anchor="e").pack(side=TOP,fill=X)
    Label(root,text="Student Registration",width=10,height=2,bg="#c36464",fg="#fff",font='arial 20 bold').pack(side=TOP,fill=X)
    
    #searchbox to update
    Search=StringVar()
    Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)
    imageicon3=Image.open("srch.png")
    imageicon3=imageicon3.resize((50,30))
    imageicon3=ImageTk.PhotoImage(imageicon3)
    srch=Button(root,text="Search",compound=LEFT,image=imageicon3,bg="#68ddfa",width=123,font="arial 13 bold",command=search)
    srch.place(x=1060,y=66)
    
    imageicon4=Image.open("update.png")
    imageicon4=imageicon4.resize((50,30))
    imageicon4=ImageTk.PhotoImage(imageicon4)
    Update_btn=Button(root,image=imageicon4,bg="#c36464",command=update)
    Update_btn.place(x=110,y=64)
    
    
    #registration and date
    Label(root,text="Enrollment No.",font="arial 13",fg=framebg,bg=bg).place(x=30,y=150)
    Label(root,text="Date",font="arial 13",fg=framebg,bg=bg).place(x=500,y=150)
    
    Registration=StringVar()
    Date = StringVar()
    
    reg_entry=Entry(root,textvariable=Registration,width=15,font="airal 10")
    reg_entry.place(x=160,y=150)
    #reg no
    
    today=date.today()
    d1=today.strftime("%d/%m/%y")
    date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")
    date_entry.place(x=550,y=150)
    
    Date.set(d1)
    
    #student details
    obj=LabelFrame(root,text="student's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=200,relief=GROOVE)
    obj.place(x=30,y=200)
    
    Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
    Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
    Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
    Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
    
    Name=StringVar()
    name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")
    name_entry.place(x=160,y=50)
    
    DOB=StringVar()
    dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")
    dob_entry.place(x=640,y=50)
    
    Class=Combobox(obj,values=['IT1','IT2','CE1','CE2','CE3'],font="Roboto 10",width=17,state="r")
    Class.set("Select Class")
    Class.place(x=640,y=100)
    
    radio=IntVar()
    R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
    R1.place(x=150,y=100)
    
    R2=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
    R2.place(x=200,y=100)
    
    
    #parents details
    obj2=LabelFrame(root,text="Parents's Details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=170,relief=GROOVE)
    obj2.place(x=30,y=410)
    
    Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)
    Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)
    
    F_Name=StringVar()
    fEntry=Entry(obj2,textvariable=F_Name,width=20,font='arial 10')
    fEntry.place(x=160,y=50)
    
    M_Name=StringVar()
    mEntry=Entry(obj2,textvariable=M_Name,width=20,font='arial 10')
    mEntry.place(x=640,y=50)
    
    
    #image
    f=Frame(root,bd=3,bg="white",width=200,height=200,relief=GROOVE)
    f.place(x=950,y=150)
    
    imageicon5=Image.open("user.jpg")
    imageicon5=imageicon5.resize((200,200))
    imageicon5=ImageTk.PhotoImage(imageicon5)
    lbl=Label(f,bg="white",image=imageicon5)
    lbl.place(x=0,y=0)
    
    #button
    Button(root,text="Upload",width=19,height=1,font="arial 12 bold",bg="lightblue",command=showimg).place(x=950,y=370)
    
    
    saveButton=Button(root,text="Save",width=19,height=1,font="arial 12 bold",bg="lightblue",command=save).place(x=950,y=430)
    
    
    Button(root,text="Reset",width=19,height=1,font="arial 12 bold",bg="lightblue",command=clear).place(x=950,y=490)
    
    
    Button(root,text="Exit",width=19,height=1,font="arial 12 bold",bg="lightblue",command=Exit).place(x=950,y=550)
    
    root.mainloop()
